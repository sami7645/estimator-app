from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Q, Max
from django.contrib.auth import get_user_model
from django.utils import timezone

import fitz  # PyMuPDF
from pathlib import Path
from io import BytesIO
import tempfile
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from openpyxl import Workbook
import math

from .models import (
    Project,
    PlanSet,
    PlanPage,
    PlanPageOverlay,
    CountDefinition,
    CountItem,
    ScaleCalibration,
    DatasetExample,
    Detection,
    ContactMessage,
    Subscription,
    SubscriptionPlan,
    SubscriptionStatus,
    TeamMember,
    TeamRole,
    PrivacyAgreement,
    TradeDataset,
    DatasetImage,
    Trade,
    ProjectEmail,
)
from .serializers import (
    ProjectSerializer,
    PlanSetSerializer,
    PlanPageSerializer,
    CountDefinitionSerializer,
    CountItemSerializer,
    ScaleCalibrationSerializer,
    DatasetExampleSerializer,
    DetectionSerializer,
    ContactMessageSerializer,
    SubscriptionSerializer,
    TeamMemberSerializer,
    PrivacyAgreementSerializer,
    TradeDatasetSerializer,
    DatasetImageSerializer,
    UserSearchSerializer,
    ProjectEmailSerializer,
)

User = get_user_model()


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.query_params.get("owner") == "me" and self.request.user.is_authenticated:
            user = self.request.user
            team_owner_ids = TeamMember.objects.filter(
                user=user, accepted=True
            ).values_list("subscription__owner_id", flat=True)
            qs = qs.filter(
                Q(owner=user) | Q(owner__isnull=True) | Q(owner_id__in=team_owner_ids)
            )

        email = self.request.query_params.get("estimating_email")
        if email:
            qs = qs.filter(estimating_email__iexact=email)

        return qs

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user if self.request.user.is_authenticated else None)


class ProjectEmailViewSet(viewsets.ModelViewSet):
    queryset = ProjectEmail.objects.all()
    serializer_class = ProjectEmailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        project = self.request.query_params.get("project")
        category = self.request.query_params.get("category")
        if project:
            qs = qs.filter(project_id=project)
        if category:
            qs = qs.filter(category=category)
        return qs


def _generate_pages_for_plan_set(plan_set: PlanSet) -> None:
    """
    Open the uploaded PDF and create PlanPage records with rendered PNGs.
    Updates plan_set.processing_pages_total / processing_pages_done for progress.
    """
    import logging
    logger = logging.getLogger(__name__)
    try:
        pdf_path = plan_set.pdf_file.path
        doc = fitz.open(pdf_path)
        media_root = Path(settings.MEDIA_ROOT)
        total = len(doc)

        plan_set.processing_pages_total = total
        plan_set.processing_pages_done = 0
        plan_set.save(update_fields=["processing_pages_total", "processing_pages_done"])

        zoom_x = 2.0
        zoom_y = 2.0
        effective_dpi_x = 72.0 * zoom_x
        effective_dpi_y = 72.0 * zoom_y

        for page_index in range(total):
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom_x, zoom_y))
            rel_dir = Path("plans") / "pages" / str(plan_set.id)
            abs_dir = media_root / rel_dir
            abs_dir.mkdir(parents=True, exist_ok=True)
            filename = f"page_{page_index + 1}.png"
            abs_path = abs_dir / filename
            pix.save(abs_path.as_posix())

            PlanPage.objects.create(
                plan_set=plan_set,
                page_number=page_index + 1,
                image=str(rel_dir / filename),
                title="",
                dpi_x=effective_dpi_x,
                dpi_y=effective_dpi_y,
            )
            plan_set.processing_pages_done = page_index + 1
            plan_set.save(update_fields=["processing_pages_done"])

        doc.close()
    except Exception as exc:
        logger.exception("PDF page generation failed for plan_set=%s: %s", plan_set.id, exc)
    finally:
        plan_set.processing_pages_total = None
        plan_set.processing_pages_done = None
        plan_set.save(update_fields=["processing_pages_total", "processing_pages_done"])


class PlanSetViewSet(viewsets.ModelViewSet):
    queryset = PlanSet.objects.all().prefetch_related("pages")
    serializer_class = PlanSetSerializer

    @transaction.atomic
    def perform_create(self, serializer):
        plan_set = serializer.save()
        import threading
        threading.Thread(target=_generate_pages_for_plan_set, args=(plan_set,), daemon=True).start()


class PlanPageViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PlanPage.objects.all()
    serializer_class = PlanPageSerializer
    parser_classes = [MultiPartParser, FormParser] + list(viewsets.ReadOnlyModelViewSet.parser_classes)

    def _save_uploaded_alt_image(self, page, file, is_pdf):
        """Save uploaded file as image. Returns (image_field_value,). For PlanPage.image_alt or PlanPageOverlay.image."""
        if is_pdf:
            raw = file.read()
            doc = fitz.open(stream=raw, filetype="pdf")
            if len(doc) == 0:
                doc.close()
                raise ValueError("PDF has no pages.")
            first_page = doc.load_page(0)
            zoom_x, zoom_y = 2.0, 2.0
            pix = first_page.get_pixmap(matrix=fitz.Matrix(zoom_x, zoom_y))
            doc.close()
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                pix.save(tmp.name)
                with open(tmp.name, "rb") as f:
                    content = ContentFile(f.read())
            Path(tmp.name).unlink(missing_ok=True)
            return content, f"page_{page.id}_alt.png"
        save_name = file.name or f"page_{page.id}_alt.png"
        if not save_name.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
            save_name = f"page_{page.id}_alt.png"
        return file, save_name

    @action(detail=True, methods=["post"])
    def upload_alt(self, request, pk=None):
        """
        Upload an extra background image for this page (e.g. satellite view).
        First upload sets image_alt; further uploads create PlanPageOverlay entries (multiple images per page).
        Accepts: image (jpg/png/jpeg etc) or PDF (first page converted to PNG).
        Body: multipart form with key "file".
        """
        page = self.get_object()
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided. Use form key 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        name_lower = (file.name or "").lower()
        is_pdf = name_lower.endswith(".pdf") or (getattr(file, "content_type", "") or "").lower() == "application/pdf"

        try:
            content_or_file, save_name = self._save_uploaded_alt_image(page, file, is_pdf)
            def next_auto_name():
                existing = []
                if page.image_alt_name:
                    existing.append(page.image_alt_name.strip().lower())
                for o in page.overlays.all():
                    if o.name:
                        existing.append(o.name.strip().lower())
                i = 1
                while f"image {i}".lower() in existing:
                    i += 1
                return f"Image {i}"

            if not page.image_alt:
                # First extra image: set image_alt
                page.image_alt.save(save_name, content_or_file, save=True)
                if not page.image_alt_name:
                    page.image_alt_name = next_auto_name()
                    page.save(update_fields=["image_alt_name"])
            else:
                # Already have image_alt; add as overlay (multiple images)
                max_order = (
                    page.overlays.aggregate(Max("order"))["order__max"]
                    or -1
                )
                overlay = PlanPageOverlay(plan_page=page, order=max_order + 1, name=next_auto_name())
                overlay.save()
                overlay.image.save(save_name, content_or_file, save=True)
            page.refresh_from_db()
            return Response(PlanPageSerializer(page).data)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["patch"], url_path=r"extra-images/(?P<extra_id>[^/.]+)/rename")
    def rename_extra_image(self, request, pk=None, extra_id=None):
        page = self.get_object()
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"detail": "name is required"}, status=status.HTTP_400_BAD_REQUEST)
        # Enforce uniqueness per page (case-insensitive)
        existing = []
        if page.image_alt and page.image_alt_name:
            existing.append(page.image_alt_name.strip().lower())
        for o in page.overlays.all():
            if o.name:
                existing.append(o.name.strip().lower())
        target_lower = name.lower()
        # remove current name from existing before checking
        if extra_id == "alt" and page.image_alt_name:
            try:
                existing.remove(page.image_alt_name.strip().lower())
            except ValueError:
                pass
        if extra_id not in (None, "alt"):
            try:
                o = page.overlays.get(id=int(extra_id))
                if o.name:
                    try:
                        existing.remove(o.name.strip().lower())
                    except ValueError:
                        pass
            except Exception:
                pass
        if target_lower in existing:
            return Response({"detail": "Name already exists on this sheet."}, status=status.HTTP_400_BAD_REQUEST)

        if extra_id == "alt":
            if not page.image_alt:
                return Response({"detail": "No alt image on this page."}, status=status.HTTP_400_BAD_REQUEST)
            page.image_alt_name = name
            page.save(update_fields=["image_alt_name"])
        else:
            try:
                overlay = page.overlays.get(id=int(extra_id))
            except Exception:
                return Response({"detail": "Overlay not found."}, status=status.HTTP_404_NOT_FOUND)
            overlay.name = name
            overlay.save(update_fields=["name"])
        page.refresh_from_db()
        return Response(PlanPageSerializer(page).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path=r"extra-images/(?P<extra_id>[^/.]+)/transform")
    def update_extra_image_transform(self, request, pk=None, extra_id=None):
        """Update scale / offset for an extra background image."""
        page = self.get_object()
        scale = request.data.get("scale")
        offset_x = request.data.get("offset_x")
        offset_y = request.data.get("offset_y")

        if extra_id == "alt":
            if not page.image_alt:
                return Response({"detail": "No alt image on this page."}, status=status.HTTP_400_BAD_REQUEST)
            fields = []
            if scale is not None:
                page.image_alt_scale = float(scale)
                fields.append("image_alt_scale")
            if offset_x is not None:
                page.image_alt_offset_x = float(offset_x)
                fields.append("image_alt_offset_x")
            if offset_y is not None:
                page.image_alt_offset_y = float(offset_y)
                fields.append("image_alt_offset_y")
            if fields:
                page.save(update_fields=fields)
        else:
            try:
                overlay = page.overlays.get(id=int(extra_id))
            except Exception:
                return Response({"detail": "Overlay not found."}, status=status.HTTP_404_NOT_FOUND)
            if scale is not None:
                overlay.scale = float(scale)
            if offset_x is not None:
                overlay.offset_x = float(offset_x)
            if offset_y is not None:
                overlay.offset_y = float(offset_y)
            overlay.save()
        page.refresh_from_db()
        return Response(PlanPageSerializer(page).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["delete"], url_path=r"extra-images/(?P<extra_id>[^/.]+)")
    def delete_extra_image(self, request, pk=None, extra_id=None):
        page = self.get_object()
        if extra_id == "alt":
            if not page.image_alt:
                return Response({"detail": "No alt image on this page."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                page.image_alt.delete(save=False)
            except Exception:
                pass
            page.image_alt = None
            page.image_alt_name = ""
            page.save(update_fields=["image_alt", "image_alt_name"])
        else:
            try:
                overlay = page.overlays.get(id=int(extra_id))
            except Exception:
                return Response({"detail": "Overlay not found."}, status=status.HTTP_404_NOT_FOUND)
            try:
                overlay.image.delete(save=False)
            except Exception:
                pass
            overlay.delete()
        page.refresh_from_db()
        return Response(PlanPageSerializer(page).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def add_to_dataset(self, request, pk=None):
        """
        Simple Add Page -> Dataset implementation.
        For now, we just create a DatasetExample that references the full page image
        and copies over basic metadata; cropping to individual symbols can be added later.
        Expected body: {"trade": "...", "count_definition_id": optional}
        """
        page = self.get_object()
        trade = request.data.get("trade")
        count_def_id = request.data.get("count_definition_id")
        if not trade:
            return Response({"detail": "trade is required"}, status=status.HTTP_400_BAD_REQUEST)

        rel_dir = Path("datasets") / "examples" / str(page.id)
        media_root = Path(settings.MEDIA_ROOT)
        abs_dir = media_root / rel_dir
        abs_dir.mkdir(parents=True, exist_ok=True)

        # Copy the page image into dataset folder
        src_path = media_root / Path(page.image.name)
        dst_path = abs_dir / "page_full.png"
        if src_path.is_file():
            dst_path.write_bytes(src_path.read_bytes())

        example = DatasetExample.objects.create(
            plan_page=page,
            trade=trade,
            count_definition_id=count_def_id,
            image=str(rel_dir / "page_full.png"),
            metadata={"source": "add_to_dataset_full_page"},
        )
        return Response(DatasetExampleSerializer(example).data, status=status.HTTP_201_CREATED)


class CountDefinitionViewSet(viewsets.ModelViewSet):
    queryset = CountDefinition.objects.all()
    serializer_class = CountDefinitionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        plan_set = self.request.query_params.get("plan_set")
        if plan_set:
            qs = qs.filter(plan_set_id=plan_set)
        return qs

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def upload_shape_image(self, request):
        """
        Upload an icon image for an EACH count definition.
        Accepts multipart form-data with key "file" (JPEG/PNG/etc).
        Returns: {"url": "<relative media path>"} which should be stored in shape_image_url.
        """
        file = request.FILES.get("file") or request.FILES.get("image")
        if not file:
            return Response(
                {"detail": "No file provided. Use form key 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        original_name = file.name or "shape.png"
        base_name = original_name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        if "." not in base_name:
            base_name = f"{base_name}.png"
        name_lower = base_name.lower()
        if not name_lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
            base_name = f"{base_name}.png"

        path = default_storage.save(f"count_shapes/{base_name}", file)
        # Return relative media path; frontend will prefix with MEDIA_BASE if needed.
        return Response({"url": path}, status=status.HTTP_201_CREATED)


class CountItemViewSet(viewsets.ModelViewSet):
    queryset = CountItem.objects.all()
    serializer_class = CountItemSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        cd = self.request.query_params.get("count_definition")
        page = self.request.query_params.get("page")
        plan_set = self.request.query_params.get("plan_set")
        if cd:
            qs = qs.filter(count_definition_id=cd)
        if page:
            qs = qs.filter(page_id=page)
        if plan_set:
            qs = qs.filter(page__plan_set_id=plan_set)
        return qs


class ScaleCalibrationViewSet(viewsets.ModelViewSet):
    queryset = ScaleCalibration.objects.all()
    serializer_class = ScaleCalibrationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        page = self.request.query_params.get("page")
        if page:
            qs = qs.filter(page_id=page)
        return qs

    def create(self, request, *args, **kwargs):
        """Create or update: if a calibration already exists for this page, update it."""
        page_id = request.data.get("page")
        if page_id:
            existing = ScaleCalibration.objects.filter(page_id=page_id).first()
            if existing:
                serializer = self.get_serializer(existing, data=request.data, partial=True)
                serializer.is_valid(raise_exception=True)
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)


class DatasetExampleViewSet(viewsets.ModelViewSet):
    queryset = DatasetExample.objects.all()
    serializer_class = DatasetExampleSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        plan_page = self.request.query_params.get("plan_page")
        trade = self.request.query_params.get("trade")
        if plan_page:
            qs = qs.filter(plan_page_id=plan_page)
        if trade:
            qs = qs.filter(trade=trade)
        return qs


class DetectionViewSet(viewsets.ModelViewSet):
    queryset = Detection.objects.all()
    serializer_class = DetectionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        plan_page = self.request.query_params.get("plan_page")
        trade = self.request.query_params.get("trade")
        if plan_page:
            qs = qs.filter(plan_page_id=plan_page)
        if trade:
            qs = qs.filter(trade=trade)
        return qs

    @action(detail=False, methods=["post"])
    def run_auto_detect(self, request, *args, **kwargs):
        """
        Real ML / Computer-Vision detection engine.

        Uses OpenCV (template matching, ORB feature matching, Hough line
        detection, contour analysis) to scan the target floor-plan image
        and find elements that match the *explicit dataset* of annotated
        example pages/images, not the current project’s own counts.

        More pages/images added to the dataset → smarter, more accurate
        detections, while corrections on the current job never pollute
        the training data.

        Expected input: { plan_page_id, trades: [...] }
        Returns: { items_created, definitions_created, removed_ids, count,
                   dataset_summary }
        """
        from .ml_engine import run_ml_detection

        TRADE_COLORS = {
            "acoustic": "#6366f1",
            "electrical": "#f59e0b",
            "plumbing": "#3b82f6",
            "mechanical": "#10b981",
            "other": "#8b5cf6",
        }
        COUNT_TYPE_LABELS = {
            "area_perimeter": "Area",
            "linear_feet": "Linear",
            "each": "Count",
        }

        page_id = request.data.get("plan_page_id")
        trades = request.data.get("trades", [])
        if not page_id:
            return Response(
                {"detail": "plan_page_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not trades:
            return Response(
                {"detail": "trades list is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            page = PlanPage.objects.get(id=page_id)
        except PlanPage.DoesNotExist:
            return Response({"detail": "Page not found"}, status=status.HTTP_404_NOT_FOUND)

        plan_set = page.plan_set
        items_created = []
        definitions_created = []

        # ── Step 1: Remove previous auto-detected items for these trades ──
        old_auto = CountItem.objects.filter(
            page=page,
            is_auto_detected=True,
            count_definition__trade__in=trades,
        )
        removed_ids = list(old_auto.values_list("id", flat=True))
        old_auto.delete()

        # ── Step 2: Run ML detection per trade ──
        for trade in trades:
            trade_label = trade.replace("_", " ").title()

            # Dataset: all count items (manual or previously auto) that live
            # on pages which have been explicitly added to the dataset for
            # this trade. If a page was added to the dataset for multiple
            # count definitions, every definition on that page participates,
            # but we still keep them fully separate via their own profiles.
            per_def_cap = 30
            global_items = []

            dataset_page_ids = list(
                DatasetExample.objects.filter(trade=trade)
                .values_list("plan_page_id", flat=True)
            )

            if dataset_page_ids:
                base_qs = (
                    CountItem.objects.filter(
                        page_id__in=dataset_page_ids,
                        count_definition__trade=trade,
                    )
                    .exclude(page=page)
                    .select_related("count_definition", "page")
                )
            else:
                base_qs = CountItem.objects.none()

            # Per-definition sampling gives every dataset count type a chance
            # and stops common ones from drowning out rarer ones.
            defs_for_trade = (
                CountDefinition.objects.filter(
                    id__in=base_qs.values("count_definition_id"),
                    trade=trade,
                )
                .distinct()
            )
            for cdef in defs_for_trade:
                qs = base_qs.filter(count_definition=cdef).order_by("-id")[:per_def_cap]
                global_items.extend(list(qs))

            if not global_items:
                continue

            # ── 2a: Run the CV/ML detection engine ──
            try:
                detections, profiles = run_ml_detection(
                    target_page=page,
                    trade=trade,
                    global_items=global_items,
                    max_results=45,
                )
            except Exception as exc:
                import logging
                logging.getLogger(__name__).error(
                    "ML detection failed for trade=%s page=%s: %s",
                    trade, page_id, exc,
                )
                detections, profiles = [], []

            # If there is absolutely no usable dataset for this trade, skip.
            if not detections and not profiles:
                continue

            # Build quick lookup of dataset profiles per (count_type, color, shape)
            profile_by_key: dict[tuple[str, str, str], object] = {}
            for p in profiles:
                c = p.color or TRADE_COLORS.get(trade, "#8b5cf6")
                profile_by_key[(p.count_type, c, p.shape or "")] = p

            # Ensure geometry we save is always plain Python floats,
            # never numpy float32, so JSONField can serialize it.
            def _normalize_geom(geometry):
                norm = []
                if not geometry:
                    return norm
                for pt in geometry:
                    if isinstance(pt, (list, tuple)) and len(pt) >= 2:
                        norm.append([float(pt[0]), float(pt[1])])
                    else:
                        norm.append(pt)
                return norm

            # 2b: Map every dataset CountDefinition (DefProfile) to its own
            # local CountDefinition in this plan set, keyed by the *dataset*
            # def_id. This guarantees that different dataset "count types"
            # are never merged together, even if they share color/shape.
            local_def_by_src: dict[int, CountDefinition] = {}
            type_counter: dict[str, int] = {}

            for p in profiles:
                ct = p.count_type
                color = p.color or TRADE_COLORS.get(trade, "#8b5cf6")
                shape = p.shape or ""

                # Try to reuse a definition that already matches this dataset
                # profile in this plan set (same trade, type, color, shape,
                # and name). If none exists, create a new one.
                existing = CountDefinition.objects.filter(
                    plan_set=plan_set,
                    trade=trade,
                    count_type=ct,
                    color=color,
                    shape=shape,
                    name=p.name,
                ).first()

                if existing:
                    local_def = existing
                else:
                    type_counter[ct] = type_counter.get(ct, 0) + 1
                    base_name = p.name or COUNT_TYPE_LABELS.get(ct, ct.title())
                    # If we already created a def for this count_type with
                    # the same name, append a numeric suffix to keep them
                    # distinct in the sidebar.
                    suffix = ""
                    if any(
                        d.name == base_name and d.count_type == ct and d.trade == trade
                        for d in definitions_created
                    ):
                        suffix = f" {type_counter[ct]}"
                    auto_name = base_name + suffix

                    local_def = CountDefinition.objects.create(
                        plan_set=plan_set,
                        name=auto_name,
                        trade=trade,
                        count_type=ct,
                        color=color,
                        shape=shape,
                    )
                    definitions_created.append(local_def)

                local_def_by_src[p.def_id] = local_def

            # 2b‑2: Create CountItems for actual detections, routing each
            # detection back to the CountDefinition corresponding to the
            # dataset def_id that produced it.
            profiles_by_id = {p.def_id: p for p in profiles}
            for det in detections:
                ct = det.count_type
                src_id = getattr(det, "source_def_id", None)

                local_def = None
                if src_id is not None:
                    local_def = local_def_by_src.get(src_id)

                # Fallback for older detections without source_def_id:
                if local_def is None:
                    color = det.source_color or TRADE_COLORS.get(trade, "#8b5cf6")
                    det_shape = det.source_shape or ""
                    local_def = CountDefinition.objects.filter(
                        plan_set=plan_set,
                        trade=trade,
                        count_type=ct,
                        color=color,
                        shape=det_shape,
                    ).first()
                    if not local_def:
                        type_counter[ct] = type_counter.get(ct, 0) + 1
                        label = COUNT_TYPE_LABELS.get(ct, ct.title())
                        suffix = f" {type_counter[ct]}" if type_counter[ct] > 1 else ""
                        auto_name = f"{trade_label} - {label}{suffix}"
                        local_def = CountDefinition.objects.create(
                            plan_set=plan_set,
                            name=auto_name,
                            trade=trade,
                            count_type=ct,
                            color=color,
                            shape=det_shape,
                        )
                        definitions_created.append(local_def)
                # "each" items are always stored as a point so the
                # frontend renders them with a uniform marker size
                # (same as manually placed each items). Only rotation
                # varies to match dataset patterns.
                geom_type = det.geometry_type
                geom = det.geometry
                rot = det.rotation_deg

                if ct == "each" and local_def.shape:
                    p = None
                    if src_id is not None:
                        p = profiles_by_id.get(src_id)

                    if det.geometry and det.geometry_type == "point":
                        cx, cy = det.geometry[0]

                        seed = 0
                        if p is not None:
                            seed = int(
                                (cx * 9973.0 + cy * 7919.0 + float(p.def_id) * 101.0)
                                * 10000.0
                            )

                        if p and getattr(p, "rotations_deg", None) and p.rotations_deg:
                            rs = sorted(p.rotations_deg)
                            rot = rs[abs(seed) % len(rs)]
                        else:
                            rot = 0.0

                        # Keep as point so the frontend renders it
                        # with the standard markerSize (same as manual).
                        geom_type = "point"
                        geom = [[cx, cy]]

                item = CountItem.objects.create(
                    count_definition=local_def,
                    page=page,
                    geometry_type=geom_type,
                    geometry=_normalize_geom(geom),
                    area_sqft=det.area_sqft,
                    perimeter_ft=det.perimeter_ft,
                    length_ft=det.length_ft,
                    rotation_deg=rot,
                    is_auto_detected=True,
                )
                items_created.append(item)

        # ── Step 3: Serialize response ──
        items_data = CountItemSerializer(items_created, many=True).data
        defs_data = CountDefinitionSerializer(definitions_created, many=True).data

        dataset_summary = {}
        for trade in trades:
            dataset_page_ids = list(
                DatasetExample.objects.filter(trade=trade)
                .values_list("plan_page_id", flat=True)
            )
            if dataset_page_ids:
                global_count = CountItem.objects.filter(
                    count_definition__trade=trade,
                    page_id__in=dataset_page_ids,
                ).exclude(page=page).count()
            else:
                global_count = 0

            dataset_summary[trade] = {
                "global_items": global_count,
                "total": global_count,
            }

        return Response({
            "items_created": items_data,
            "definitions_created": defs_data,
            "removed_ids": removed_ids,
            "count": len(items_created),
            "dataset_summary": dataset_summary,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def export_counts_excel(self, request, *args, **kwargs):
        """
        Export a professional Excel workbook with multiple sheets:
        1) Summary  – project info + totals per count definition
        2) Details  – every individual item with measurements
        3) By Page  – per-page breakdown grouped by count definition
        """
        from openpyxl.styles import Font, Alignment, PatternFill, numbers

        plan_set_id = request.query_params.get("plan_set")
        price_per_each = request.query_params.get("price_per_each")
        price_per_sqft = request.query_params.get("price_per_sqft")
        price_per_perimeter_ft = request.query_params.get("price_per_perimeter_ft")
        price_per_linear_ft = request.query_params.get("price_per_linear_ft")

        def _float_or_none(val):
            if val is None or val == "":
                return None
            try:
                return float(val)
            except (TypeError, ValueError):
                return None

        price_each = _float_or_none(price_per_each)
        price_sqft = _float_or_none(price_per_sqft)
        price_perimeter = _float_or_none(price_per_perimeter_ft)
        price_linear = _float_or_none(price_per_linear_ft)
        use_prices = any(x is not None for x in (price_each, price_sqft, price_perimeter, price_linear))

        defs_qs = CountDefinition.objects.select_related("plan_set", "plan_set__project")
        items_qs = CountItem.objects.select_related(
            "count_definition", "page", "count_definition__plan_set"
        ).order_by("count_definition__name", "page__page_number", "id")

        if plan_set_id:
            defs_qs = defs_qs.filter(plan_set_id=plan_set_id)
            items_qs = items_qs.filter(count_definition__plan_set_id=plan_set_id)

        plan_set = (
            PlanSet.objects.filter(id=plan_set_id)
            .select_related("project")
            .first()
            if plan_set_id
            else None
        )

        # Style helpers
        hdr_font = Font(bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        num_fmt_2 = numbers.FORMAT_NUMBER_00
        title_font = Font(bold=True, size=12)

        def style_header_row(ws, row_num=1, col_count=8):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=c)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align

        # Pages that have scale calibration (measurements only valid for these)
        calibrated_page_ids = set()
        if plan_set_id:
            calibrated_page_ids = set(
                ScaleCalibration.objects.filter(
                    page__plan_set_id=plan_set_id
                ).values_list("page_id", flat=True)
            )

        # ──────────── Sheet 1: Summary ────────────
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"

        row = 1
        if plan_set:
            ws1.cell(row=row, column=1, value="Project").font = title_font
            ws1.cell(row=row, column=2, value=plan_set.project.name if plan_set.project else "")
            row += 1
            ws1.cell(row=row, column=1, value="Plan Set").font = title_font
            ws1.cell(row=row, column=2, value=plan_set.name)
            row += 1
            if plan_set.created_at:
                ws1.cell(row=row, column=1, value="Date").font = title_font
                ws1.cell(row=row, column=2, value=plan_set.created_at.strftime("%Y-%m-%d"))
                row += 1

            # Scale calibration info per page
            scales = ScaleCalibration.objects.filter(
                page__plan_set_id=plan_set_id
            ).select_related("page")
            row += 1
            ws1.cell(row=row, column=1, value="Scale Calibrations").font = title_font
            row += 1
            if scales.exists():
                for sc in scales:
                    ws1.cell(row=row, column=1, value=f"Page {sc.page.page_number}")
                    ppi = sc.real_world_feet / sc.pixel_distance if sc.pixel_distance else 0
                    ws1.cell(row=row, column=2, value=f"1 px = {ppi:.4f} ft")
                    row += 1
            else:
                ws1.cell(row=row, column=1, value="(None)")
                ws1.cell(row=row, column=2, value="No scale set. Calibrate in the app for real measurements.")
                row += 1

            row += 1

        # Summary table header
        headers = [
            "Count Name", "Type", "Trade", "Color",
            "Total Items", "Total Area (sqft)", "Total Perimeter (ft)", "Total Length (ft)",
        ]
        if use_prices:
            headers.append("Total Price")
        for ci, h in enumerate(headers, 1):
            ws1.cell(row=row, column=ci, value=h)
        style_header_row(ws1, row, len(headers))
        row += 1

        for cd in defs_qs:
            items = items_qs.filter(count_definition=cd)
            n = items.count()
            calibrated_items = [i for i in items if i.page_id in calibrated_page_ids]
            a = sum(i.area_sqft or 0 for i in calibrated_items)
            p = sum(i.perimeter_ft or 0 for i in calibrated_items)
            l = sum(i.length_ft or 0 for i in calibrated_items)
            row_data = [
                cd.name,
                cd.get_count_type_display(),
                cd.get_trade_display(),
                cd.color,
                n,
                round(a, 2) if a else "",
                round(p, 2) if p else "",
                round(l, 2) if l else "",
            ]
            if use_prices:
                total_price = (n or 0) * (price_each or 0) + (a or 0) * (price_sqft or 0) + (p or 0) * (price_perimeter or 0) + (l or 0) * (price_linear or 0)
                row_data.append(round(total_price, 2) if total_price else "")
            ws1.append(row_data)

        col_widths = [("A", 25), ("B", 18), ("C", 15), ("D", 12),
                      ("E", 12), ("F", 18), ("G", 18), ("H", 18)]
        if use_prices:
            col_widths.append(("I", 14))
        for w, width in col_widths:
            ws1.column_dimensions[w].width = width

        # ──────────── Sheet 2: Details ────────────
        ws2 = wb.create_sheet("Details")
        detail_hdrs = [
            "#", "Count Name", "Type", "Trade", "Page #", "Page Title",
            "Geometry", "Area (sqft)", "Perimeter (ft)", "Length (ft)", "Vertices", "Rotation (deg)",
        ]
        if use_prices:
            detail_hdrs.append("Price")
        for ci, h in enumerate(detail_hdrs, 1):
            ws2.cell(row=1, column=ci, value=h)
        style_header_row(ws2, 1, len(detail_hdrs))

        row_n = 0
        for item in items_qs:
            row_n += 1
            cd = item.count_definition
            verts = len(item.geometry) if isinstance(item.geometry, list) else 0
            # Only show real-world measurements for calibrated pages
            page_calibrated = item.page_id in calibrated_page_ids
            area_val = round(item.area_sqft, 2) if page_calibrated and item.area_sqft else 0
            perim_val = round(item.perimeter_ft, 2) if page_calibrated and item.perimeter_ft else 0
            length_val = round(item.length_ft, 2) if page_calibrated and item.length_ft else 0
            row_data = [
                row_n,
                cd.name if cd else "",
                cd.get_count_type_display() if cd else "",
                cd.get_trade_display() if cd else "",
                item.page.page_number if item.page else "",
                item.page.title or (f"Sheet {item.page.page_number}" if item.page else ""),
                item.geometry_type,
                area_val if page_calibrated and item.area_sqft else "",
                perim_val if page_calibrated and item.perimeter_ft else "",
                length_val if page_calibrated and item.length_ft else "",
                verts,
                round(item.rotation_deg, 1) if item.rotation_deg else "",
            ]
            if use_prices:
                item_price = 1 * (price_each or 0) + (area_val or 0) * (price_sqft or 0) + (perim_val or 0) * (price_perimeter or 0) + (length_val or 0) * (price_linear or 0)
                row_data.append(round(item_price, 2) if item_price else "")
            ws2.append(row_data)

        detail_cols = [("A", 6), ("B", 25), ("C", 18), ("D", 15), ("E", 10),
                       ("F", 18), ("G", 12), ("H", 15), ("I", 15), ("J", 15), ("K", 10), ("L", 14)]
        if use_prices:
            detail_cols.append(("M", 14))
        for w, width in detail_cols:
            ws2.column_dimensions[w].width = width

        # Number formatting for measurement and price columns
        price_col = 13 if use_prices else None
        for r in range(2, ws2.max_row + 1):
            for c in [8, 9, 10] + ([price_col] if price_col else []):
                cell = ws2.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt_2

        # ──────────── Sheet 3: By Page ────────────
        ws3 = wb.create_sheet("By Page")
        page_hdrs = [
            "Page #", "Page Title", "Count Name", "Type", "Trade",
            "Items", "Total Area (sqft)", "Total Perimeter (ft)", "Total Length (ft)",
        ]
        if use_prices:
            page_hdrs.append("Total Price")
        for ci, h in enumerate(page_hdrs, 1):
            ws3.cell(row=1, column=ci, value=h)
        style_header_row(ws3, 1, len(page_hdrs))

        pages_qs = PlanPage.objects.filter(plan_set_id=plan_set_id).order_by("page_number") if plan_set_id else PlanPage.objects.all()
        for page in pages_qs:
            page_items = items_qs.filter(page=page)
            if not page_items.exists():
                continue
            cd_ids = page_items.values_list("count_definition", flat=True).distinct()
            for cd_id in cd_ids:
                try:
                    cd = CountDefinition.objects.get(id=cd_id)
                except CountDefinition.DoesNotExist:
                    continue
                cd_items = page_items.filter(count_definition=cd)
                page_calibrated = page.id in calibrated_page_ids
                n = cd_items.count()
                a = sum(i.area_sqft or 0 for i in cd_items) if page_calibrated else 0
                p = sum(i.perimeter_ft or 0 for i in cd_items) if page_calibrated else 0
                l = sum(i.length_ft or 0 for i in cd_items) if page_calibrated else 0
                row_data = [
                    page.page_number,
                    page.title or f"Sheet {page.page_number}",
                    cd.name,
                    cd.get_count_type_display(),
                    cd.get_trade_display(),
                    n,
                    round(a, 2) or "" if page_calibrated else "",
                    round(p, 2) or "" if page_calibrated else "",
                    round(l, 2) or "" if page_calibrated else "",
                ]
                if use_prices:
                    total_price = (n or 0) * (price_each or 0) + (a or 0) * (price_sqft or 0) + (p or 0) * (price_perimeter or 0) + (l or 0) * (price_linear or 0)
                    row_data.append(round(total_price, 2) if total_price else "")
                ws3.append(row_data)

        by_page_cols = [("A", 10), ("B", 20), ("C", 25), ("D", 18), ("E", 15),
                        ("F", 10), ("G", 18), ("H", 18), ("I", 18)]
        if use_prices:
            by_page_cols.append(("J", 14))
        for w, width in by_page_cols:
            ws3.column_dimensions[w].width = width

        # ── Write response ──
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"counts_{plan_set.name if plan_set else 'all'}.xlsx"
        safe_filename = filename.replace('"', "'")
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}"'
        return response


# ──────────────────────────────────────────────────────────────
#  Subscription & Team Management Views
# ──────────────────────────────────────────────────────────────

class TradeDatasetViewSet(viewsets.ModelViewSet):
    queryset = TradeDataset.objects.all()
    serializer_class = TradeDatasetSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.is_authenticated:
            qs = qs.filter(owner=self.request.user)
        trade = self.request.query_params.get("trade")
        if trade:
            qs = qs.filter(trade=trade)
        return qs

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    @action(detail=True, methods=["post"])
    def upload_image(self, request, pk=None):
        dataset = self.get_object()
        image = request.FILES.get("image")
        label = request.data.get("label", "")
        if not image:
            return Response({"detail": "image file is required"}, status=status.HTTP_400_BAD_REQUEST)

        di = DatasetImage.objects.create(
            dataset=dataset,
            image=image,
            label=label,
        )
        return Response(DatasetImageSerializer(di).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def stats(self, request, pk=None):
        dataset = self.get_object()
        uploaded = dataset.images.count()
        page_examples = DatasetExample.objects.filter(
            trade=dataset.trade,
            plan_page__plan_set__project__owner=dataset.owner,
        ).count()
        return Response({
            "trade": dataset.trade,
            "uploaded_images": uploaded,
            "page_examples": page_examples,
            "total": uploaded + page_examples,
            "estimated_accuracy": min(95, 30 + ((uploaded + page_examples) * 5)),
        })


class DatasetImageViewSet(viewsets.ModelViewSet):
    queryset = DatasetImage.objects.all()
    serializer_class = DatasetImageSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        dataset = self.request.query_params.get("dataset")
        if dataset:
            qs = qs.filter(dataset_id=dataset)
        return qs


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def subscription_detail(request):
    sub, created = Subscription.objects.get_or_create(
        owner=request.user,
        defaults={"plan": SubscriptionPlan.FREE, "status": SubscriptionStatus.ACTIVE},
    )
    serializer = SubscriptionSerializer(sub)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def subscription_create_or_update(request):
    plan = request.data.get("plan", SubscriptionPlan.PRO)
    sub, created = Subscription.objects.get_or_create(
        owner=request.user,
        defaults={
            "plan": plan,
            "status": SubscriptionStatus.ACTIVE,
            "current_period_start": timezone.now(),
            "current_period_end": timezone.now() + timezone.timedelta(days=30),
        },
    )
    if not created:
        sub.plan = plan
        sub.status = SubscriptionStatus.ACTIVE
        sub.current_period_start = timezone.now()
        sub.current_period_end = timezone.now() + timezone.timedelta(days=30)
        sub.save()
    return Response(SubscriptionSerializer(sub).data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def subscription_skip(request):
    """Simulate successful subscription (skip payment for testing)."""
    plan = request.data.get("plan", SubscriptionPlan.PRO)
    sub, created = Subscription.objects.get_or_create(
        owner=request.user,
        defaults={
            "plan": plan,
            "status": SubscriptionStatus.ACTIVE,
            "current_period_start": timezone.now(),
            "current_period_end": timezone.now() + timezone.timedelta(days=30),
        },
    )
    if not created:
        sub.plan = plan
        sub.status = SubscriptionStatus.ACTIVE
        sub.current_period_start = timezone.now()
        sub.current_period_end = timezone.now() + timezone.timedelta(days=30)
        sub.save()
    return Response(SubscriptionSerializer(sub).data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def subscription_cancel(request):
    try:
        sub = Subscription.objects.get(owner=request.user)
        sub.status = SubscriptionStatus.CANCELLED
        sub.save()
        return Response(SubscriptionSerializer(sub).data)
    except Subscription.DoesNotExist:
        return Response({"detail": "No subscription found"}, status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def team_list(request):
    try:
        sub = Subscription.objects.get(owner=request.user)
    except Subscription.DoesNotExist:
        return Response({"detail": "No subscription found"}, status=status.HTTP_404_NOT_FOUND)
    members = TeamMember.objects.filter(subscription=sub).select_related("user")
    return Response(TeamMemberSerializer(members, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def team_add_member(request):
    username_or_email = request.data.get("username_or_email", "").strip()
    role = request.data.get("role", TeamRole.VIEWER)
    if not username_or_email:
        return Response({"detail": "username_or_email is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        sub = Subscription.objects.get(owner=request.user)
    except Subscription.DoesNotExist:
        return Response({"detail": "No subscription found"}, status=status.HTTP_404_NOT_FOUND)

    current_count = sub.team_members.count()
    if current_count >= sub.max_team_members:
        return Response(
            {"detail": f"Team is full ({sub.max_team_members} members max)"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    target_user = User.objects.filter(
        Q(username=username_or_email) | Q(email=username_or_email)
    ).first()

    if not target_user:
        return Response({"detail": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    if target_user == request.user:
        return Response({"detail": "Cannot add yourself"}, status=status.HTTP_400_BAD_REQUEST)

    member, created = TeamMember.objects.get_or_create(
        subscription=sub,
        user=target_user,
        defaults={"role": role, "invited_email": target_user.email or "", "accepted": True},
    )
    if not created:
        member.role = role
        member.save()

    return Response(TeamMemberSerializer(member).data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def team_update_member(request, member_id):
    try:
        sub = Subscription.objects.get(owner=request.user)
        member = TeamMember.objects.get(id=member_id, subscription=sub)
    except (Subscription.DoesNotExist, TeamMember.DoesNotExist):
        return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

    role = request.data.get("role")
    if role:
        member.role = role
        member.save()
    return Response(TeamMemberSerializer(member).data)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def team_remove_member(request, member_id):
    try:
        sub = Subscription.objects.get(owner=request.user)
        member = TeamMember.objects.get(id=member_id, subscription=sub)
    except (Subscription.DoesNotExist, TeamMember.DoesNotExist):
        return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

    member.delete()
    return Response({"detail": "Member removed"}, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def privacy_agree(request):
    agreement, created = PrivacyAgreement.objects.get_or_create(
        user=request.user,
        defaults={
            "agreed": True,
            "agreed_at": timezone.now(),
            "ip_address": request.META.get("REMOTE_ADDR"),
        },
    )
    if not created:
        agreement.agreed = True
        agreement.agreed_at = timezone.now()
        agreement.ip_address = request.META.get("REMOTE_ADDR")
        agreement.save()
    return Response(PrivacyAgreementSerializer(agreement).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def privacy_status(request):
    try:
        agreement = PrivacyAgreement.objects.get(user=request.user)
        return Response(PrivacyAgreementSerializer(agreement).data)
    except PrivacyAgreement.DoesNotExist:
        return Response({"agreed": False})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_search(request):
    q = request.query_params.get("q", "").strip()
    if len(q) < 2:
        return Response([])
    users = User.objects.filter(
        Q(username__icontains=q) | Q(email__icontains=q)
    ).exclude(id=request.user.id)[:10]
    return Response(UserSearchSerializer(users, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dataset_stats_all(request):
    """Return dataset stats for all trades for the current user.

    Optional query params:
      plan_set_id  – if provided, also returns ``items_in_planset``
                     (manual count items on OTHER pages in that plan set)
                     and ``current_page_id`` excludes that page.
      page_id      – the current page to exclude from the planset count.
    """
    plan_set_id = request.query_params.get("plan_set_id")
    page_id = request.query_params.get("page_id")

    stats = {}
    for trade_val, trade_label in Trade.choices:
        # Global dataset: manual items on pages that have been
        # explicitly saved to the dataset for this trade.
        dataset_page_ids = list(
            DatasetExample.objects.filter(trade=trade_val)
            .values_list("plan_page_id", flat=True)
        )
        if dataset_page_ids:
            global_qs = CountItem.objects.filter(
                count_definition__trade=trade_val,
                is_auto_detected=False,
                page_id__in=dataset_page_ids,
            )
        else:
            global_qs = CountItem.objects.none()
        if page_id:
            global_qs = global_qs.exclude(page_id=page_id)
        global_count = global_qs.count()

        items_in_planset = 0
        if plan_set_id:
            qs = CountItem.objects.filter(
                count_definition__plan_set_id=plan_set_id,
                count_definition__trade=trade_val,
                is_auto_detected=False,
            )
            if page_id:
                qs = qs.exclude(page_id=page_id)
            items_in_planset = qs.count()

        total = global_count
        accuracy = min(98, 20 + int(15 * math.log2(max(1, total)))) if total > 0 else 0
        stats[trade_val] = {
            "label": trade_label,
            "global_items": global_count,
            "items_in_planset": items_in_planset,
            "total": total,
            "estimated_accuracy": accuracy,
        }
    return Response(stats)


# ──────────────────────────────────────────────────────────────
#  Contact & Health
# ──────────────────────────────────────────────────────────────

@api_view(["POST"])
def contact_create(request):
    """Submit contact form. If authenticated, link message to user."""
    serializer = ContactMessageSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = request.user if request.user.is_authenticated else None
    serializer.save(user=user)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
def contact_list(request):
    """List contact messages for the current user (to view replies). Auth required."""
    if not request.user.is_authenticated:
        return Response({"detail": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
    messages = ContactMessage.objects.filter(user=request.user).order_by("-created_at")
    serializer = ContactMessageSerializer(messages, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def health_check(_request):
    return Response({"status": "ok"})

